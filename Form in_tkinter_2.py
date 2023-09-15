from tkinter import*
import pandas as pd
from openpyxl import *

wb = load_workbook('C:\\Ishan\\python\\info.xlsx')

def btnClick(inp):
    global res
    if inp == True:
        res = True
    else:
        res = False
    return res

def focus1(event): 
    global e1
    e1.focus_set()

def focus2(event): 
    global e2
    e2.focus_set()

def focus3(event): 
    global e3 
    e3.focus_set()

def insert(): 
      
    # if user not fill any entry 
    # then print "empty input" 
    if (e1.get() == "" and
        e2.get() == "" and
        e3.get() == "" and): 
              
        print("empty input") 
  
    else: 
  
        # assigning the max row and max column 
        # value upto which data is written 
        # in an excel sheet to the variable 
        current_row = sheet.max_row 
        current_column = sheet.max_column 
  
        # get method returns current text 
        # as string which we write into 
        # excel spreadsheet at particular location 
        sheet.cell(row=current_row + 1, column=1).value = e1.get() 
        sheet.cell(row=current_row + 1, column=2).value = e2.get() 
        sheet.cell(row=current_row + 1, column=3).value = e3.get() 

        # save the file 
        wb.save('C:\\Files\\Ishan\\Python\\Info.xlsx') 
  
        # set focus on the name_field box 
        name_field.focus_set() 
  
        # call the clear() function 
        clear() 

top = Tk()  
top.geometry("400x250")
top.title("Form")
name = Label(top, text = "Name").place(x = 30,y = 50)  
Class = Label(top, text = "Class").place(x = 30, y = 90)  
Roll_no = Label(top, text = "Roll No.").place(x = 30, y = 130)  
e1 = Entry(top).place(x = 80, y = 50)  
e2 = Entry(top).place(x = 80, y = 90)  
e3 = Entry(top).place(x = 95, y = 130)
submit = Button(top, text="Submit", fg="Black",  bg="Red", command=btnClick(True))
submit.place(x = 130, y = 170)


if res == True:
    d = {'Name':[e1], 'Class':[e2], 'Roll no':[e3]}
    re = pd.DataFrame(d)
    Ex = pd.ExcelWriter('Info.xlsx', engine = 'xlsxwriter')
    re.to_excel(Ex, sheet_name = 'Sheet1')
    Ex.save()

top.mainloop()
