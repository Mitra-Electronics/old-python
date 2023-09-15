# import openpyxl and tkinter modules 
from openpyxl import *
from tkinter import *

# globally declare wb and sheet variable 

# opening the existing excel file 
wb = load_workbook('C:\\Files\\Ishan\\Python\\Info2.xlsx') 

# create the sheet object 
sheet = wb.active 


def excel(): 
	
	# resize the width of columns in 
	# excel spreadsheet 
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 10
	sheet.column_dimensions['C'].width = 10
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 40
	sheet.column_dimensions['G'].width = 50

	# write given data to an excel spreadsheet 
	# at particular location 
	sheet.cell(row=1, column=1).value = "Name"
	sheet.cell(row=1, column=2).value = "Class"
	sheet.cell(row=1, column=3).value = "Roll no"


# Function to set focus (cursor) 
def focus1(event): 
	# set focus on the course_field box 
	Class_field.focus_set() 


# Function to set focus 
def focus2(event): 
	# set focus on the sem_field box 
	Roll_no_field.focus_set()

# Function to set focus 
def focus3(event): 
	# set focus on the sem_field box 
	Name_field.focus_set()

# Function for clearing the 
# contents of text entry boxes 
def clear(): 
	
	# clear the content of text entry box 
	Name_field.delete(0, END) 
	Class_field.delete(0, END) 
	Roll_no_field.delete(0, END)


# Function to take data from GUI 
# window and write to an excel file 
def insert(): 
	
	# if user not fill any entry 
	# then print "empty input" 
	if (Name_field.get() == "" and
		Class_field.get() == int and
		Roll_no_field.get() == "") : 
			
		print("empty input") 

	else: 

		# assigning the max row and max column 
		# value upto which data is written 
		# in an excel sheet to the variable 
		current_row = sheet.max_row 
		current_column = sheet.max_column 

		# get method returns current text 
		# as string which we write into 
		# excel spreadsheet at particular location 
		sheet.cell(row=current_row + 1, column=1).value = Name_field.get() 
		sheet.cell(row=current_row + 1, column=2).value = Class_field.get() 
		sheet.cell(row=current_row + 1, column=3).value = Roll_no_field.get()

		# save the file 
		wb.save('C:\\Files\\Ishan\\Python\\Info2.xlsx') 

		# set focus on the name_field box 
		Name_field.focus_set() 

		# call the clear() function 
		clear() 


# Driver code 
if __name__ == "__main__": 
	
	# create a GUI window 
	root = Tk() 

	# set the background colour of GUI window 
	root.configure(background='light green') 

	# set the title of GUI window 
	root.title("registration form") 

	# set the configuration of GUI window 
	root.geometry("500x300") 

	excel() 

	# create a Form label 
	heading = Label(root, text="Form", bg="light green") 

	# create a Name label 
	Name = Label(root, text="Name", bg="light green") 

	# create a Course label 
	Class = Label(root, text="Class", bg="light green") 

	# create a Roll_no label 
	Roll_no = Label(root, text="Roll_no", bg="light green") 

	# grid method is used for placing 
	# the widgets at respective positions 
	# in table like structure . 
	heading.grid(row=0, column=1) 
	Name.grid(row=1, column=0) 
	Class.grid(row=2, column=0) 
	Roll_no.grid(row=3, column=0)

	# create a text entry box 
	# for typing the information 
	Name_field = Entry(root) 
	Class_field = Entry(root) 
	Roll_no_field = Entry(root)  

	# bind method of widget is used for 
	# the binding the function with the events 

	# whenever the enter key is pressed 
	# then call the focus1 function 
	Name_field.bind("<Ret`urn>", focus1) 

	# whenever the enter key is pressed 
	# then call the focus2 function 
	Class_field.bind("<Return>", focus2) 

	# whenever the enter key is pressed 
	# then call the focus3 function 
	Roll_no_field.bind("<Return>", focus3) 

	# grid method is used for placing 
	# the widgets at respective positions 
	# in table like structure . 
	Name_field.grid(row=1, column=1, ipadx="100") 
	Class_field.grid(row=2, column=1, ipadx="100") 
	Roll_no_field.grid(row=3, column=1, ipadx="100") 

	# call excel function 
	excel() 

	# create a Submit Button and place into the root window 
	submit = Button(root, text="Submit", fg="Black", 
							bg="Red", command=insert) 
	submit.grid(row=8, column=1) 

	# start the GUI 
	root.mainloop() 
